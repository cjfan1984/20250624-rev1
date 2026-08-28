from __future__ import annotations

import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


MASTER_SHEET = "三级SKU主库"
SNAPSHOT_SHEET = "SKU决策快照"

MASTER_HEADERS = [
    "一级系统", "二级产品族", "三级SKU", "数据依据", "首选履约", "当前优先级", "备注",
    "俄文关键词", "采购关键词", "供货平台", "首选供货链接", "备选供货链接", "采购价CNY",
    "MOQ", "毛重kg", "包装长cm", "包装宽cm", "包装高cm", "计费重kg", "WB售价RUB",
    "OZON售价RUB", "跨境净利润CNY", "本土净利润CNY", "上架顺序", "研究状态",
    "证据置信度", "更新时间", "补充说明", "WB源数据", "OZON源数据", "源表索引",
]

SNAPSHOT_REQUIRED_HEADERS = [
    "SKU_KEY", "主库行", "产品名称/SKU", "当前阶段", "单件净利润CNY", "净利率",
    "快照更新时间", "Ozon有效竞品链接", "WB有效竞品链接",
]

DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
HTTP_RE = re.compile(r"^https?://", re.I)
SHEET_INDEX_RE = re.compile(r"^[^!\r\n]+![A-Z]+\d+(?::[A-Z]+\d+)?$", re.I)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _is_iso_date(value: Any) -> bool:
    text = _text(value)
    if not DATE_RE.fullmatch(text):
        return False
    try:
        date.fromisoformat(text)
    except ValueError:
        return False
    return True


def _issue(row: int, code: str, message: str, severity: str = "high") -> dict[str, Any]:
    return {"row": row, "severity": severity, "code": code, "message": message}


def validate_headers(actual: Iterable[Any], required: Iterable[str], *, exact: bool = False) -> None:
    headers = [_text(x) for x in actual]
    expected = list(required)
    if exact and headers[: len(expected)] != expected:
        raise ValueError("sheet header contract drifted")
    missing = [name for name in expected if name not in headers]
    if missing:
        raise ValueError(f"sheet missing required headers: {missing}")


def validate_master_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    """Detect corruption that commonly indicates a shifted Google Sheets write.

    Blank evidence is allowed. Values in a populated row must remain in their
    typed columns; the validator intentionally does not invent or repair data.
    """
    issues: list[dict[str, Any]] = []
    names: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        sku = _text(row.get("三级SKU"))
        if not sku:
            continue
        names.append(sku)

        updated = _text(row.get("更新时间"))
        if updated and not _is_iso_date(updated):
            issues.append(_issue(row_number, "MASTER_UPDATED_AT_INVALID", "更新时间不是 YYYY-MM-DD；疑似字段错位"))

        wb_source = _text(row.get("WB源数据"))
        ozon_source = _text(row.get("OZON源数据"))
        source_index = _text(row.get("源表索引"))
        if "ozon.ru/" in wb_source.lower():
            issues.append(_issue(row_number, "WB_SOURCE_HAS_OZON_URL", "WB源数据列出现 Ozon URL；疑似字段错位"))
        if "wildberries.ru/" in ozon_source.lower():
            issues.append(_issue(row_number, "OZON_SOURCE_HAS_WB_URL", "OZON源数据列出现 WB URL；疑似字段错位"))
        if source_index and HTTP_RE.match(source_index):
            issues.append(_issue(row_number, "SOURCE_INDEX_IS_URL", "源表索引列出现 URL；疑似字段错位"))
        elif source_index and not (
            SHEET_INDEX_RE.fullmatch(source_index)
            or source_index.startswith(("R", "证据", "每日新品", "SKU证据库"))
        ):
            issues.append(_issue(row_number, "SOURCE_INDEX_UNUSUAL", "源表索引格式异常", "medium"))

        master_row = row.get("上架顺序")
        if _text(master_row) and not re.fullmatch(r"\d+(?:\.0+)?", _text(master_row)):
            issues.append(_issue(row_number, "MASTER_ORDER_INVALID", "上架顺序不是数字", "medium"))

    duplicates = sorted(name for name, count in Counter(names).items() if count > 1)
    for name in duplicates[:20]:
        issues.append(_issue(0, "MASTER_DUPLICATE_SKU", f"重复三级SKU：{name}"))
    high = [item for item in issues if item["severity"] == "high"]
    return {
        "sheet": MASTER_SHEET,
        "rows": len(names),
        "uniqueSkus": len(set(names)),
        "issueCount": len(issues),
        "highRiskCount": len(high),
        "highRiskRows": sorted({item["row"] for item in high if item["row"]}),
        "issues": issues,
        "status": "blocked" if high else ("review" if issues else "ok"),
    }


def validate_snapshot_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    issues: list[dict[str, Any]] = []
    keys: list[str] = []
    names: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        key = _text(row.get("SKU_KEY"))
        name = _text(row.get("产品名称/SKU"))
        if not key and not name:
            continue
        if not key or not name:
            issues.append(_issue(row_number, "SNAPSHOT_IDENTITY_MISSING", "SKU_KEY 或产品名称缺失"))
            continue
        keys.append(key)
        names.append(name)
        updated = _text(row.get("快照更新时间"))
        if updated and not _is_iso_date(updated):
            issues.append(_issue(row_number, "SNAPSHOT_UPDATED_AT_INVALID", "快照更新时间不是 YYYY-MM-DD；疑似字段错位"))
        if "|" not in key:
            issues.append(_issue(row_number, "SKU_KEY_INVALID", "SKU_KEY 缺少编号与名称分隔符", "medium"))

    for key, count in Counter(keys).items():
        if count > 1:
            issues.append(_issue(0, "SNAPSHOT_DUPLICATE_KEY", f"重复 SKU_KEY：{key}"))
    for name, count in Counter(names).items():
        if count > 1:
            issues.append(_issue(0, "SNAPSHOT_DUPLICATE_NAME", f"重复产品名称：{name}"))
    high = [item for item in issues if item["severity"] == "high"]
    return {
        "sheet": SNAPSHOT_SHEET,
        "rows": len(keys),
        "uniqueSkuKeys": len(set(keys)),
        "issueCount": len(issues),
        "highRiskCount": len(high),
        "highRiskRows": sorted({item["row"] for item in high if item["row"]}),
        "issues": issues,
        "status": "blocked" if high else ("review" if issues else "ok"),
    }


def _rows_from_values(values: list[list[Any]], *, exact_headers: list[str] | None = None) -> list[dict[str, Any]]:
    if not values:
        raise ValueError("sheet is empty")
    headers = [_text(x) for x in values[0]]
    if exact_headers is not None:
        validate_headers(headers, exact_headers, exact=True)
    rows = []
    for values_row in values[1:]:
        rows.append({headers[i]: values_row[i] if i < len(values_row) else None for i in range(len(headers))})
    return rows


def load_master_rows_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if MASTER_SHEET not in wb.sheetnames:
        raise ValueError(f"worksheet not found: {MASTER_SHEET}")
    values = [list(row) for row in wb[MASTER_SHEET].iter_rows(values_only=True)]
    return _rows_from_values(values, exact_headers=MASTER_HEADERS)

def load_master_rows_google(spreadsheet_id: str, credentials: Any) -> list[dict[str, Any]]:
    import gspread

    gc = gspread.authorize(credentials)
    values = gc.open_by_key(spreadsheet_id).worksheet(MASTER_SHEET).get_all_values()
    return _rows_from_values(values, exact_headers=MASTER_HEADERS)
