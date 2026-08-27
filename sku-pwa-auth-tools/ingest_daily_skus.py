from __future__ import annotations
import argparse,json,os,re,sys
from pathlib import Path
HERE=Path(__file__).resolve().parent;sys.path.insert(0,str(HERE))
from automation_modules import ingest_plan,text,is_direct_product_url

MASTER='三级SKU主库'; DAILY='每日新品_决策卡数据层'
MASTER_HEADERS=['一级系统','二级产品族','三级SKU','数据依据','首选履约','当前优先级','备注','俄文关键词','采购关键词','供货平台','首选供货链接','备选供货链接','采购价CNY','MOQ','毛重kg','包装长cm','包装宽cm','包装高cm','计费重kg','WB售价RUB','OZON售价RUB','跨境净利润CNY','本土净利润CNY','上架顺序','研究状态','证据置信度','更新时间','补充说明','WB源数据','OZON源数据','源表索引']

def n(v):
    if v is None:return None
    m=re.search(r'-?\d+(?:\.\d+)?',str(v).replace(',',''))
    return float(m.group()) if m else None

def load_xlsx(path:Path):
    from openpyxl import load_workbook
    wb=load_workbook(path,read_only=True,data_only=True)
    m=wb[MASTER]; mh=[str(x).strip() if x is not None else '' for x in next(m.iter_rows(values_only=True))];mi=mh.index('三级SKU');master=[text(r[mi]) for r in m.iter_rows(min_row=2,values_only=True) if len(r)>mi and text(r[mi])]
    d=wb[DAILY]; rows=list(d.iter_rows(values_only=True));hi=next(i for i,r in enumerate(rows) if r and '产品/型号/规格' in r);headers=[str(x).strip() if x is not None else '' for x in rows[hi]];out=[]
    for sheet_row,vals in enumerate(rows[hi+1:],start=hi+2):
        row={headers[i]:vals[i] if i<len(vals) else None for i in range(len(headers))}
        if text(row.get('产品/型号/规格')):row['_sheet_row']=sheet_row;out.append(row)
    return master,out

def google_creds(write=False):
    from google.oauth2.service_account import Credentials
    raw=os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON');fn=os.getenv('GOOGLE_SERVICE_ACCOUNT_FILE');scopes=['https://www.googleapis.com/auth/spreadsheets' if write else 'https://www.googleapis.com/auth/spreadsheets.readonly','https://www.googleapis.com/auth/drive.readonly']
    if raw:return Credentials.from_service_account_info(json.loads(raw),scopes=scopes)
    if fn:return Credentials.from_service_account_file(fn,scopes=scopes)
    raise SystemExit('Google credentials missing')

def load_google(spreadsheet_id:str,write=False):
    import gspread
    gc=gspread.authorize(google_creds(write));sh=gc.open_by_key(spreadsheet_id);mws=sh.worksheet(MASTER);master=[text(v) for v in mws.col_values(3)[1:] if text(v)];dws=sh.worksheet(DAILY);rows=dws.get_all_values();hi=next(i for i,r in enumerate(rows) if '产品/型号/规格' in r);headers=rows[hi];out=[]
    for sheet_row,vals in enumerate(rows[hi+1:],start=hi+2):
        row={headers[i]:vals[i] if i<len(vals) else None for i in range(len(headers))}
        if text(row.get('产品/型号/规格')):row['_sheet_row']=sheet_row;out.append(row)
    return master,out,sh

def build_plan(master:list[str],candidates:list[dict]):
    already=[];unlinked=[]
    for c in candidates:
        if n(c.get('主库行')):
            already.append({'candidate':text(c.get('产品/型号/规格')),'masterRow':int(n(c.get('主库行'))),'status':'已有主库行'})
        else:unlinked.append(c)
    dedupe=ingest_plan([text(c.get('产品/型号/规格')) for c in unlinked],master)
    return {'linkedExisting':already,**dedupe,'counts':{'linkedExisting':len(already),**dedupe['counts']},'_unlinked':unlinked}

def candidate_to_master_row(c:dict)->list:
    supplier=text(c.get('货源链接')) if is_direct_product_url(c.get('货源链接')) else None
    confidence=(text(c.get('证据完整度')) or '')[:60] or '待补'
    return [
      text(c.get('一级系统')),text(c.get('产品族')),text(c.get('产品/型号/规格')),
      f"每日新品自动入库｜{text(c.get('数据粒度')) or '候选'}",'待核',text(c.get('当前优先级')) or 'C',text(c.get('当前动作')),
      None,text(c.get('产品/型号/规格')),text(c.get('货源平台/供应商')),supplier,None,None,n(c.get('MOQ')),n(c.get('毛重kg')),n(c.get('包装长cm')),n(c.get('包装宽cm')),n(c.get('包装高cm')),n(c.get('计费重kg')),None,n(c.get('前台当前价RUB')),None,None,None,text(c.get('利润状态')) or '候选/待补',confidence,text(c.get('最近研究日期')) or text(c.get('最近证据日期')),text(c.get('缺失字段')),text(c.get('WB证据')),text(c.get('Ozon商品卡')),f"{DAILY}!{c.get('_sheet_row')}"
    ]

def apply_new(spreadsheet_id:str,plan:dict)->list[dict]:
    import gspread
    master,candidates,sh=load_google(spreadsheet_id,write=True);mws=sh.worksheet(MASTER);dws=sh.worksheet(DAILY);master_header=mws.row_values(1)
    if master_header[:len(MASTER_HEADERS)]!=MASTER_HEADERS:raise SystemExit('master header contract drifted; refusing auto-ingest')
    daily_headers=dws.row_values(3);master_row_col=daily_headers.index('主库行')+1
    by_name={text(c.get('产品/型号/规格')):c for c in candidates}
    existing_col=mws.col_values(3);next_row=max([i for i,v in enumerate(existing_col,start=1) if text(v)] or [1])+1
    results=[]
    for item in plan.get('newSku',[]):
        name=item['candidate'];c=by_name.get(name)
        if not c:continue
        values=candidate_to_master_row(c)
        mws.update(range_name=f'A{next_row}:AE{next_row}',values=[values],value_input_option='USER_ENTERED')
        dws.update_cell(int(c['_sheet_row']),master_row_col,next_row)
        results.append({'sku':name,'masterRow':next_row});next_row+=1
    return results

def public_plan(plan:dict)->dict:return {k:v for k,v in plan.items() if not k.startswith('_')}

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--xlsx',type=Path);ap.add_argument('--spreadsheet-id',default='108o5gtxkUsWEZI8xfZFE89Kq83IfStQGNm6dvjIlAkk');ap.add_argument('--output',type=Path,default=Path('ingest-plan.json'));ap.add_argument('--apply',action='store_true');a=ap.parse_args()
    if a.xlsx:
        if a.apply:raise SystemExit('--apply requires Google API mode')
        master,candidates=load_xlsx(a.xlsx)
    else:
        master,candidates,_=load_google(a.spreadsheet_id,write=a.apply)
    plan=build_plan(master,candidates);plan['masterCount']=len(master);plan['candidateCount']=len(candidates);plan['policy']='已有主库行直接合并；未关联候选>=0.985自动判重，0.80–0.985人工复核，<0.80才允许自动建新SKU'
    if a.apply:
        plan['applied']=apply_new(a.spreadsheet_id,plan)
        plan['masterCountAfter']=len(master)+len(plan['applied'])
    a.output.write_text(json.dumps(public_plan(plan),ensure_ascii=False,indent=2),encoding='utf-8');print(json.dumps(public_plan(plan)['counts'],ensure_ascii=False))
if __name__=='__main__':main()
