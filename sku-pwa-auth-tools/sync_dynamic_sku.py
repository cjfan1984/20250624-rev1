from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

DEFAULT_SPREADSHEET_ID = "108o5gtxkUsWEZI8xfZFE89Kq83IfStQGNm6dvjIlAkk"
DEFAULT_WORKSHEET = "SKU决策快照"
DEFAULT_MASTER_WORKSHEET = "三级SKU主库"
PENDING_TEXT = ("待补", "未闭环", "待核", "待研究", "待询价", "缺价", "缺采购", "待建模")


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def as_number(value: Any) -> float | int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return int(value) if isinstance(value, float) and value.is_integer() else value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    is_percent = text.endswith("%")
    text = text.rstrip("%").replace("¥", "").replace("RUB", "").strip()
    try:
        number = float(text)
    except ValueError:
        return None
    if is_percent:
        number /= 100
    return int(number) if number.is_integer() else number


def is_pending(value: Any) -> bool:
    text = clean_text(value)
    return bool(text and any(token in text for token in PENDING_TEXT))


def is_http_url(value: Any) -> bool:
    text = clean_text(value)
    return bool(text and re.match(r"^https?://", text, flags=re.I))


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (list, tuple)):
        return [json_safe(v) for v in value]
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    return str(value)


def canonical_json(value: Any) -> str:
    return json.dumps(json_safe(value), ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def sha256_text(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def load_from_xlsx(path: Path, worksheet: str) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if worksheet not in wb.sheetnames:
        raise SystemExit(f"worksheet not found: {worksheet}")
    ws = wb[worksheet]
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise SystemExit("worksheet is empty")
    headers = [clean_text(x) or f"__col_{i+1}" for i, x in enumerate(header_row)]
    out: list[dict[str, Any]] = []
    for values in rows:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if clean_text(row.get("SKU_KEY")) or clean_text(row.get("产品名称/SKU")):
            out.append(row)
    return out


def _google_credentials():
    from google.oauth2.service_account import Credentials
    raw_json = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    cred_file = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE")
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
    if raw_json:
        return Credentials.from_service_account_info(json.loads(raw_json), scopes=scopes)
    if cred_file:
        return Credentials.from_service_account_file(cred_file, scopes=scopes)
    raise SystemExit("Google credentials missing: set GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_SERVICE_ACCOUNT_FILE")


def load_from_google(spreadsheet_id: str, worksheet: str) -> list[dict[str, Any]]:
    import gspread
    gc = gspread.authorize(_google_credentials())
    ws = gc.open_by_key(spreadsheet_id).worksheet(worksheet)
    records = ws.get_all_records(default_blank=None, numericise_ignore=["all"])
    return [dict(row) for row in records if clean_text(row.get("SKU_KEY")) or clean_text(row.get("产品名称/SKU"))]


def load_master_names_from_xlsx(path: Path, worksheet: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    if worksheet not in wb.sheetnames:
        raise SystemExit(f"master worksheet not found: {worksheet}")
    rows = wb[worksheet].iter_rows(values_only=True)
    try:
        header = [clean_text(x) for x in next(rows)]
    except StopIteration:
        raise SystemExit("master worksheet is empty")
    try:
        sku_index = header.index("三级SKU")
    except ValueError as exc:
        raise SystemExit("master worksheet missing 三级SKU column") from exc
    names = [clean_text(row[sku_index]) for row in rows if sku_index < len(row)]
    return [name for name in names if name]


def load_master_names_from_google(spreadsheet_id: str, worksheet: str) -> list[str]:
    import gspread
    gc = gspread.authorize(_google_credentials())
    ws = gc.open_by_key(spreadsheet_id).worksheet(worksheet)
    header = ws.row_values(1)
    try:
        col = header.index("三级SKU") + 1
    except ValueError as exc:
        raise SystemExit("master worksheet missing 三级SKU column") from exc
    return [name for name in (clean_text(v) for v in ws.col_values(col)[1:]) if name]


def validate_master_coverage(master_names: list[str], source_rows: list[dict[str, Any]]) -> None:
    if len(master_names) != len(set(master_names)):
        raise SystemExit("duplicate 三级SKU names found in master library")
    snapshot_names = [clean_text(row.get("产品名称/SKU")) for row in source_rows]
    snapshot_names = [name for name in snapshot_names if name]
    if len(snapshot_names) != len(set(snapshot_names)):
        raise SystemExit("duplicate 产品名称/SKU names found in decision snapshot")
    master_set = set(master_names)
    snapshot_set = set(snapshot_names)
    if master_set != snapshot_set:
        missing = sorted(master_set - snapshot_set)[:8]
        extra = sorted(snapshot_set - master_set)[:8]
        raise SystemExit(
            f"dynamic SKU coverage mismatch: master={len(master_set)} snapshot={len(snapshot_set)} "
            f"missing_in_snapshot={missing} extra_in_snapshot={extra}"
        )


def source_row_to_contract(source: dict[str, Any], ordinal: int) -> dict[str, Any]:
    sku_key = clean_text(source.get("SKU_KEY"))
    sku = clean_text(source.get("产品名称/SKU")) or (sku_key.split("|", 1)[1] if sku_key and "|" in sku_key else sku_key)
    if not sku_key or not sku:
        raise ValueError("SKU_KEY and SKU name are required")
    stage = clean_text(source.get("当前阶段")) or "CANDIDATE_PENDING"
    profit = as_number(source.get("单件净利润CNY"))
    margin = as_number(source.get("净利率"))
    if margin is not None and margin > 1.5:
        margin = margin / 100
    cost = as_number(source.get("实际/核价采购成本CNY"))
    price = as_number(source.get("Ozon目标/当前售价RUB"))
    weight = as_number(source.get("毛重kg"))
    cost_status = clean_text(source.get("采购价状态"))
    freight = clean_text(source.get("跨境物流"))
    platform_fee = clean_text(source.get("平台费金额/费率"))
    profit_known = profit is not None and margin is not None
    formal_like = stage in {"PROFIT_GATE_OK", "READY", "READY_TO_LIST"}
    if stage == "STOP":
        pwa_tier = "STOP_AUDIT"
    elif formal_like:
        pwa_tier = "FORMAL_PROFIT"
    elif profit_known:
        pwa_tier = "MODEL_PROFIT"
    else:
        pwa_tier = "CANDIDATE_PENDING"
    profit_rank_eligible = bool(
        formal_like and profit_known and not is_pending(cost_status) and not is_pending(freight) and not is_pending(platform_fee)
    )
    or_pass = bool(profit_known and ((profit is not None and profit > 10) or (margin is not None and margin > 0.15)))
    current_sources = [
        clean_text(source.get("候选货源1")),
        clean_text(source.get("候选货源2")),
        clean_text(source.get("候选货源3")),
    ]
    current_sources = [x for x in current_sources if is_http_url(x)]
    normalized_source = json_safe({k: v for k, v in source.items() if k and not str(k).startswith("__col_")})
    row_hash = sha256_text(canonical_json(normalized_source))
    return {
        "rank": ordinal,
        "sourceRank": as_number(source.get("决策排名")),
        "skuKey": sku_key,
        "sku": sku,
        "stage": stage,
        "priority": clean_text(source.get("当前优先级")),
        "pwaTier": pwa_tier,
        "profitKnown": profit_known,
        "profitRankEligible": profit_rank_eligible,
        "decisionOrPass": or_pass,
        "profit": profit,
        "margin": margin,
        "price": price,
        "priceKnown": price is not None,
        "cost": cost,
        "costKnown": cost is not None,
        "weight": weight,
        "weightKnown": weight is not None,
        "localWbProfit": as_number(source.get("本土WB模型利润CNY")),
        "localOzonProfit": as_number(source.get("本土Ozon模型利润CNY")),
        "localBestProfit": as_number(source.get("本土最佳模型利润CNY")),
        "localConservativeProfit": as_number(source.get("本土保守模型利润CNY")),
        "localWbRank": as_number(source.get("WB本土利润排名")),
        "localOzonRank": as_number(source.get("Ozon本土利润排名")),
        "localOverallRank": as_number(source.get("本土综合排名")),
        "supplierUrls": current_sources,
        "ozonCompetitorUrl": clean_text(source.get("Ozon有效竞品链接")) if is_http_url(source.get("Ozon有效竞品链接")) else None,
        "ozonCompetitorMatch": clean_text(source.get("Ozon竞品匹配层级")),
        "wbCompetitorUrl": clean_text(source.get("WB有效竞品链接")) if is_http_url(source.get("WB有效竞品链接")) else None,
        "wbCompetitorMatch": clean_text(source.get("WB竞品匹配层级")),
        "imageUrl": clean_text(source.get("PWA参考图URL")),
        "imageSourceType": clean_text(source.get("PWA图片来源类型")),
        "imageSourcePage": clean_text(source.get("图片源页面")),
        "rowHash": row_hash,
        "source": normalized_source,
    }


def build_dataset(source_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    staged: list[tuple[float, int, dict[str, Any]]] = []
    for index, source in enumerate(source_rows, start=1):
        sku_key = clean_text(source.get("SKU_KEY"))
        if not sku_key:
            continue
        if sku_key in seen:
            raise SystemExit(f"duplicate SKU_KEY: {sku_key}")
        seen.add(sku_key)
        source_rank = as_number(source.get("决策排名"))
        sort_rank = float(source_rank) if isinstance(source_rank, (int, float)) else float("inf")
        staged.append((sort_rank, index, source))
    staged.sort(key=lambda item: (item[0], item[1]))
    records = [source_row_to_contract(source, ordinal=i) for i, (_, _, source) in enumerate(staged, start=1)]
    if not records:
        raise SystemExit("no SKU records found")
    return records


def load_previous(path: Path | None) -> dict[str, str]:
    if not path or not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        return {}
    return {str(r.get("skuKey")): str(r.get("rowHash")) for r in data if r.get("skuKey") and r.get("rowHash")}


def make_manifest(records: list[dict[str, Any]], previous_hashes: dict[str, str]) -> dict[str, Any]:
    current_hashes = {r["skuKey"]: r["rowHash"] for r in records}
    added = sorted(set(current_hashes) - set(previous_hashes))
    removed = sorted(set(previous_hashes) - set(current_hashes))
    changed = sorted(k for k in set(current_hashes) & set(previous_hashes) if current_hashes[k] != previous_hashes[k])
    stable_projection = [{k: v for k, v in r.items() if k != "rowHash"} for r in records]
    dataset_sha = sha256_text(canonical_json(stable_projection))
    now = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    return {
        "schema": "SKU-DYNAMIC-DATA-V1",
        "generatedAt": now,
        "records": len(records),
        "datasetSha256": dataset_sha,
        "version": dataset_sha[:16],
        "delta": {"added": len(added), "changed": len(changed), "removed": len(removed)},
        "stats": {
            "stop": sum(r["stage"] == "STOP" for r in records),
            "profitKnown": sum(bool(r["profitKnown"]) for r in records),
            "orPass": sum(bool(r["decisionOrPass"]) for r in records),
            "formalProfitRankEligible": sum(bool(r["profitRankEligible"]) for r in records),
            "pending": sum(r["pwaTier"] == "CANDIDATE_PENDING" for r in records),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build dynamic SKU PWA data from Google Sheet or exported XLSX.")
    parser.add_argument("--spreadsheet-id", default=DEFAULT_SPREADSHEET_ID)
    parser.add_argument("--worksheet", default=DEFAULT_WORKSHEET)
    parser.add_argument("--master-worksheet", default=DEFAULT_MASTER_WORKSHEET)
    parser.add_argument("--xlsx", type=Path, help="Read an exported XLSX instead of Google API (safe fallback/test mode).")
    parser.add_argument("--output", type=Path, default=Path("new-pwa-data.json"))
    parser.add_argument("--manifest", type=Path, default=Path("pwa-data-version.json"))
    parser.add_argument("--previous", type=Path, help="Previous generated JSON used for per-SKU hash delta.")
    parser.add_argument("--private-delta", type=Path, help="Optional private delta report with SKU_KEYs. Never publish this file.")
    args = parser.parse_args()
    if args.xlsx:
        source_rows = load_from_xlsx(args.xlsx, args.worksheet)
        master_names = load_master_names_from_xlsx(args.xlsx, args.master_worksheet)
    else:
        source_rows = load_from_google(args.spreadsheet_id, args.worksheet)
        master_names = load_master_names_from_google(args.spreadsheet_id, args.master_worksheet)
    validate_master_coverage(master_names, source_rows)
    records = build_dataset(source_rows)
    previous_hashes = load_previous(args.previous)
    manifest = make_manifest(records, previous_hashes)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(records, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    args.manifest.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.private_delta:
        current_hashes = {r["skuKey"]: r["rowHash"] for r in records}
        added = sorted(set(current_hashes) - set(previous_hashes))
        removed = sorted(set(previous_hashes) - set(current_hashes))
        changed = sorted(k for k in set(current_hashes) & set(previous_hashes) if current_hashes[k] != previous_hashes[k])
        args.private_delta.parent.mkdir(parents=True, exist_ok=True)
        args.private_delta.write_text(
            json.dumps({"added": added, "changed": changed, "removed": removed}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    print(json.dumps(manifest, ensure_ascii=False, separators=(",", ":")))


if __name__ == "__main__":
    main()
