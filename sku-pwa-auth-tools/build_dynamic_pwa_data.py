from __future__ import annotations

import argparse, hashlib, json, re, sys
from pathlib import Path
from typing import Any

HERE=Path(__file__).resolve().parent
sys.path.insert(0,str(HERE))
import sync_dynamic_sku as base
from automation_modules import MODULE_REGISTRY, CORE_BUNDLES, enrich_automation

PARAM_SHEET='全SKU_交叉透视'

def load_params_xlsx(path:Path)->dict[str,float]:
    from openpyxl import load_workbook
    wb=load_workbook(path,read_only=True,data_only=True);ws=wb[PARAM_SHEET];vals={}
    for row in ws.iter_rows(min_row=5,max_row=12,values_only=True):
        if row and len(row)>=2: vals[str(row[0])]=row[1]
    return {'rub_per_cny':float(vals.get('RUB/CNY',12.31981)),'platform_rate':float(vals.get('跨境比例费',0.215)),'fixed_fee_cny':float(vals.get('跨境固定费CNY',3.0)),'profit_threshold_cny':float(vals.get('单件净利润OR门槛CNY',10.0)),'margin_threshold':float(vals.get('跨境最低净利率',0.15)),'target_margin':float(vals.get('建议目标净利率',0.20))}

def load_params_google(spreadsheet_id:str)->dict[str,float]:
    import gspread
    gc=gspread.authorize(base._google_credentials());ws=gc.open_by_key(spreadsheet_id).worksheet(PARAM_SHEET);rows=ws.get('A5:B12');vals={r[0]:r[1] for r in rows if len(r)>=2}
    def f(k,d):
        try:return float(str(vals.get(k,d)).replace(',',''))
        except:return d
    return {'rub_per_cny':f('RUB/CNY',12.31981),'platform_rate':f('跨境比例费',.215),'fixed_fee_cny':f('跨境固定费CNY',3),'profit_threshold_cny':f('单件净利润OR门槛CNY',10),'margin_threshold':f('跨境最低净利率',.15),'target_margin':f('建议目标净利率',.20)}

def canonical(obj:Any)->str:return json.dumps(obj,ensure_ascii=False,sort_keys=True,separators=(',',':'))

def load_previous(path:Path|None)->dict[str,str]:
    if not path or not path.exists(): return {}
    try:data=json.loads(path.read_text(encoding='utf-8'))
    except:return {}
    return {str(r.get('skuKey')):str(r.get('rowHash')) for r in data if isinstance(r,dict) and r.get('skuKey') and r.get('rowHash')}

def parse_dims(value:Any)->tuple[float|None,float|None,float|None]:
    s=str(value or '').replace('×','x').lower()
    nums=re.findall(r'\d+(?:\.\d+)?',s)
    if len(nums)>=3:return tuple(float(x) for x in nums[:3])
    return (None,None,None)

def parse_first_number(value:Any)->float|None:
    if value is None:return None
    m=re.search(r'-?\d+(?:\.\d+)?',str(value).replace(',',''))
    return float(m.group()) if m else None

def legacy_compat(record:dict[str,Any])->None:
    s=record.get('source') or {}
    master_row=base.as_number(s.get('主库行'))
    if master_row is not None:record['rank']=int(master_row)
    margin_fraction=record.get('margin')
    record['marginFraction']=margin_fraction
    if isinstance(margin_fraction,(int,float)):record['margin']=round(margin_fraction*100,4)
    record['formalProfitClosed']=record.get('pwaTier')=='FORMAL_PROFIT'
    record['profitRankEligible']=bool(record.get('profitKnown') and record.get('stage')!='STOP')
    l,w,h=parse_dims(s.get('包装尺寸cm'))
    record.update({
      'system':base.clean_text(s.get('一级系统')),'family':base.clean_text(s.get('产品族')),
      'updated':base.clean_text(s.get('快照更新时间')) or base.clean_text(s.get('快照时间')),
      'researchStatus':base.clean_text(s.get('研究状态')),'officialCoverage':base.clean_text(s.get('官方需求覆盖')),
      'evidence':base.clean_text(s.get('采购价状态')),'matchLevel':base.clean_text(s.get('证据匹配层级')),
      'supplyPlatform':base.clean_text(s.get('货源平台')),'moq':base.as_number(s.get('MOQ')),
      'blocker':base.clean_text(s.get('唯一阻塞点')),'spec':base.clean_text(s.get('完整目标产品规格')),
      'reviewInsights':base.clean_text(s.get('竞品高频差评')),'reviewEvidence':base.clean_text(s.get('差评证据')),
      'improvement':base.clean_text(s.get('改进/验货要求')),'packageL':l,'packageW':w,'packageH':h,
      'chargeWeight':base.as_number(s.get('计费重kg')),'targetCost':base.as_number(s.get('15%目标采购成本上限CNY')),
      'cel':parse_first_number(s.get('跨境物流')),'ozonUnits':base.as_number(s.get('Ozon竞品/市场销量')),
      'wbOrders':base.as_number(s.get('WB竞品/市场销量')),'wbSearch':None,
      'ozForecast':[0,0],'wbForecast':[0,0],'totalForecast':[0,0],
    })
    suppliers=[u for u in [s.get('候选货源1'),s.get('候选货源2'),s.get('候选货源3')] if base.is_http_url(u)]
    oz=s.get('Ozon有效竞品链接') if base.is_http_url(s.get('Ozon有效竞品链接')) else None
    wb=s.get('WB有效竞品链接') if base.is_http_url(s.get('WB有效竞品链接')) else None
    record['links']={'1688':suppliers[0] if suppliers else None,'taobao':None,'pdd':None,'wb':wb,'ozon':oz}
    record['supplierLinks']=[{'url':u,'match':'候选货源'} for u in suppliers]
    record['linkStatus']={'supply':f'已核{len(suppliers)}家' if suppliers else '待补','wb':base.clean_text(s.get('WB竞品匹配层级')) or ('商品页' if wb else '待补'),'ozon':base.clean_text(s.get('Ozon竞品匹配层级')) or ('商品页' if oz else '待补')}

def main():
    ap=argparse.ArgumentParser();ap.add_argument('--xlsx',type=Path);ap.add_argument('--spreadsheet-id',default=base.DEFAULT_SPREADSHEET_ID);ap.add_argument('--output',type=Path,default=Path('new-pwa-data.json'));ap.add_argument('--manifest',type=Path,default=Path('pwa-data-version.json'));ap.add_argument('--previous',type=Path);ap.add_argument('--private-delta',type=Path);args=ap.parse_args()
    if args.xlsx:
        rows=base.load_from_xlsx(args.xlsx,base.DEFAULT_WORKSHEET);master=base.load_master_names_from_xlsx(args.xlsx,base.DEFAULT_MASTER_WORKSHEET);params=load_params_xlsx(args.xlsx)
    else:
        rows=base.load_from_google(args.spreadsheet_id,base.DEFAULT_WORKSHEET);master=base.load_master_names_from_google(args.spreadsheet_id,base.DEFAULT_MASTER_WORKSHEET);params=load_params_google(args.spreadsheet_id)
    base.validate_master_coverage(master,rows);records=base.build_dataset(rows)
    for record in records:
        auto=enrich_automation(record.get('source') or {},params);record['automation']=auto;record['gapCount']=auto['gaps']['gapCount'];record['completion']=auto['gaps']['completion'];record['primaryBlocker']=auto['gaps']['primaryBlocker'];record['queueLevel']=auto['queue']['level'];record['queueScore']=auto['queue']['score'];record['structuralStop']=auto['profitGate'].get('structuralStop');record['supplierCount']=auto['sourcing']['supplierCount'];legacy_compat(record);record['rowHash']=hashlib.sha256(canonical({'source':record.get('source'),'automation':auto,'params':params}).encode()).hexdigest()
    prev=load_previous(args.previous);current={r['skuKey']:r['rowHash'] for r in records};added=sorted(set(current)-set(prev));removed=sorted(set(prev)-set(current));changed=sorted(k for k in current.keys()&prev.keys() if current[k]!=prev[k]);stable=[{k:v for k,v in r.items() if k!='rowHash'} for r in records];sha=hashlib.sha256(canonical(stable).encode()).hexdigest()
    manifest={'schema':'SKU-DYNAMIC-AUTOMATION-V3','version':sha[:16],'datasetSha256':sha,'records':len(records),'delta':{'added':len(added),'changed':len(changed),'removed':len(removed)},'stats':{'stop':sum(r['stage']=='STOP' for r in records),'profitKnown':sum(bool(r.get('profitKnown')) for r in records),'orPass':sum(bool(r.get('decisionOrPass')) for r in records),'structuralStop':sum(r.get('structuralStop') is True for r in records),'modelRankEligible':sum(bool(r.get('profitRankEligible')) for r in records),'p1':sum(r.get('queueLevel')=='P1' for r in records),'missingCritical':sum(bool(r.get('primaryBlocker')) for r in records)},'coreBundles':CORE_BUNDLES,'automationModules':MODULE_REGISTRY,'params':params}
    args.output.parent.mkdir(parents=True,exist_ok=True);args.manifest.parent.mkdir(parents=True,exist_ok=True);args.output.write_text(json.dumps(records,ensure_ascii=False,separators=(',',':')),encoding='utf-8');args.manifest.write_text(json.dumps(manifest,ensure_ascii=False,indent=2),encoding='utf-8')
    if args.private_delta:
        args.private_delta.parent.mkdir(parents=True,exist_ok=True);args.private_delta.write_text(json.dumps({'added':added,'changed':changed,'removed':removed},ensure_ascii=False,indent=2),encoding='utf-8')
    print(json.dumps(manifest,ensure_ascii=False,separators=(',',':')))
if __name__=='__main__': main()
