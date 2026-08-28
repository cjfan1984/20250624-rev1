import sys
import json
from pathlib import Path


TOOLS = Path(__file__).parents[1] / "sku-pwa-auth-tools"
sys.path.insert(0, str(TOOLS))

import advanced_modules
import publish_dynamic_data
import select_new_skus
import sheet_contracts


def valid_master_row(**overrides):
    row = {header: "" for header in sheet_contracts.MASTER_HEADERS}
    row.update(
        {
            "三级SKU": "测试SKU",
            "更新时间": "2026-08-28",
            "WB源数据": "WB搜索词级证据",
            "OZON源数据": "https://www.ozon.ru/product/1234567",
            "源表索引": "SKU证据库!A2:H8",
            "上架顺序": "1",
        }
    )
    row.update(overrides)
    return row


def ozon_row(name, family, url, score=80, units=500, price=900):
    return {
        "产品大类": "工具耗材",
        "查询主题": family,
        "中文子类目": family,
        "商品名称": name,
        "商品链接": url,
        "货号": name + "-01",
        "蓝海综合分": score,
        "已订购商品_件": units,
        "平均购买价格_RUB": price,
        "CEL运费占售价": 0.3,
        "数据周期": "28天",
        "订购金额_RUB": units * price,
    }


def test_master_contract_blocks_shifted_columns():
    report = sheet_contracts.validate_master_rows(
        [
            valid_master_row(
                更新时间="一段本应位于补充说明的长文本",
                WB源数据="https://www.ozon.ru/product/1234567",
                源表索引="https://www.ozon.ru/product/1234567",
            )
        ]
    )
    assert report["status"] == "blocked"
    assert report["highRiskCount"] == 3
    assert {issue["code"] for issue in report["issues"]} == {
        "MASTER_UPDATED_AT_INVALID",
        "WB_SOURCE_HAS_OZON_URL",
        "SOURCE_INDEX_IS_URL",
    }


def test_master_contract_accepts_repaired_row():
    report = sheet_contracts.validate_master_rows([valid_master_row()])
    assert report["status"] == "ok"
    assert report["highRiskCount"] == 0


def test_snapshot_invalid_update_date_is_high_risk():
    report = sheet_contracts.validate_snapshot_rows(
        [{"SKU_KEY": "0001|测试", "产品名称/SKU": "测试", "快照更新时间": "2026/08/28"}]
    )
    assert report["highRiskCount"] == 1
    assert report["issues"][0]["code"] == "SNAPSHOT_UPDATED_AT_INVALID"


def test_advanced_audit_flags_shifted_update_date():
    report = advanced_modules.data_quality_audit({"source": {"快照更新时间": "错误列内容"}})
    assert report["highRiskCount"] == 1
    assert report["issues"][0]["code"] == "UPDATED_AT_INVALID"


def test_dataset_hash_ignores_row_hash_and_is_stable():
    records = [{"skuKey": "1|A", "value": 1, "rowHash": "random-a"}]
    first = publish_dynamic_data.dataset_digest(records)
    records[0]["rowHash"] = "random-b"
    assert publish_dynamic_data.dataset_digest(records) == first
    records[0]["value"] = 2
    assert publish_dynamic_data.dataset_digest(records) != first


def test_publish_skips_encryption_when_dataset_hash_is_unchanged(tmp_path, monkeypatch):
    from openpyxl import Workbook

    workbook_path = tmp_path / "fixture.xlsx"
    workbook = Workbook()
    master = workbook.active
    master.title = "三级SKU主库"
    master.append(sheet_contracts.MASTER_HEADERS)
    master.append([valid_master_row().get(header, "") for header in sheet_contracts.MASTER_HEADERS])

    snapshot = workbook.create_sheet("SKU决策快照")
    snapshot_headers = [
        "SKU_KEY", "产品名称/SKU", "决策排名", "当前阶段", "当前优先级", "单件净利润CNY",
        "净利率", "实际/核价采购成本CNY", "采购价状态", "Ozon目标/当前售价RUB", "毛重kg",
        "跨境物流", "平台费金额/费率", "快照更新时间", "Ozon有效竞品链接", "WB有效竞品链接",
    ]
    snapshot.append(snapshot_headers)
    snapshot.append(
        [
            "0001|测试SKU", "测试SKU", 1, "MODEL_PROFIT_OK", "A", 11, 0.16, 5,
            "正式PI", 900, 0.2, "171 RUB", "21.5%", "2026-08-28",
            "https://www.ozon.ru/product/1234567", "https://www.wildberries.ru/catalog/1234567/detail.aspx",
        ]
    )

    params = workbook.create_sheet("全SKU_交叉透视")
    parameter_values = [
        ("RUB/CNY", 12.31981),
        ("跨境比例费", 0.215),
        ("跨境固定费CNY", 3),
        ("单件净利润OR门槛CNY", 10),
        ("跨境最低净利率", 0.15),
        ("建议目标净利率", 0.20),
        ("占位1", 0),
        ("占位2", 0),
    ]
    # The production loader reads A5:B12.
    for row_number, (key, value) in enumerate(parameter_values, start=5):
        params.cell(row=row_number, column=1, value=key)
        params.cell(row=row_number, column=2, value=value)
    workbook.save(workbook_path)

    first = tmp_path / "first"
    keyring = Path(__file__).parents[1] / "sku-pwa-auth-src" / "hybrid-keyring.json"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "publish_dynamic_data.py", "--xlsx", str(workbook_path), "--keyring", str(keyring),
            "--output-dir", str(first), "--fail-contract-risk",
        ],
    )
    publish_dynamic_data.main()
    assert (first / "sku-data.enc.json").exists()

    second = tmp_path / "second"
    monkeypatch.setattr(
        sys,
        "argv",
        [
            "publish_dynamic_data.py", "--xlsx", str(workbook_path), "--keyring", str(keyring),
            "--output-dir", str(second), "--previous-manifest", str(first / "pwa-data-version.json"),
            "--skip-unchanged", "--fail-contract-risk",
        ],
    )
    publish_dynamic_data.main()
    status = json.loads((second / "pipeline-status.json").read_text(encoding="utf-8"))
    assert status["state"] == "NO_CHANGE"
    assert status["shouldPublish"] is False
    assert not (second / "sku-data.enc.json").exists()


def test_selector_is_deterministic_deduped_and_family_diverse():
    ozon = [
        ozon_row("A1", "砂轮", "https://www.ozon.ru/product/a-1000001"),
        ozon_row("A2", "砂轮", "https://www.ozon.ru/product/a-1000002", score=90),
        ozon_row("B1", "测量工具", "https://www.ozon.ru/product/b-1000003", score=85),
        ozon_row("低分", "其他", "https://www.ozon.ru/product/c-1000004", score=60),
    ]
    wb = [
        {"中文产品": "砂轮", "综合分": 90},
        {"中文产品": "测量工具", "综合分": 80},
    ]
    selected = select_new_skus.select_candidates(
        ozon,
        wb,
        existing_names=[],
        existing_urls=["https://www.ozon.ru/product/b-1000003"],
        existing_ids=[],
        max_new=5,
    )
    assert len(selected) == 1
    assert selected[0]["ozon"]["商品名称"] == "A2"


def test_candidate_row_keeps_model_weight_out_of_fact_columns():
    candidate = {
        "ozon": ozon_row("A", "砂轮", "https://www.ozon.ru/product/a-1000001"),
        "wb": None,
        "score": 82,
        "productId": "1000001",
    }
    candidate["ozon"].update({"包装重量估算_kg": 0.25, "包装长估算_cm": 20})
    row = select_new_skus.candidate_to_daily(candidate, "2026-08-28")
    assert row["毛重kg"] == ""
    assert row["包装长cm"] == ""
    assert "模型值不写入事实列" in row["尺寸等级"]
